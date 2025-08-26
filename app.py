from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import io
import re

def limpar_nome_planilha(nome):
    """Remove caracteres inválidos para nomes de planilhas Excel"""
    # Caracteres inválidos: \ / ? * [ ]
    nome_limpo = re.sub(r'[\\/?*[\]]', '', nome)
    # Substituir espaços múltiplos por um único espaço
    nome_limpo = re.sub(r'\s+', ' ', nome_limpo)
    # Limitar a 31 caracteres (limite do Excel)
    return nome_limpo[:31].strip()

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///estoque.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Modelos do banco de dados
class Produto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    quantidade = db.Column(db.Integer, default=0)
    valor_custo = db.Column(db.Float, default=0.0)
    valor_varejo = db.Column(db.Float, default=0.0)  # Novo campo para preço de venda simples
    valor_atacado = db.Column(db.Float, default=0.0) # Novo campo para preço de atacadista
    compatibilidade = db.Column(db.Text)
    data_cadastro = db.Column(db.DateTime, default=datetime.utcnow)

class Venda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_venda = db.Column(db.DateTime, default=datetime.utcnow)
    valor_total = db.Column(db.Float, default=0.0)
    parcelas = db.Column(db.Integer, default=1)
    cliente = db.Column(db.String(200))
    produtos_vendidos = db.relationship('ItemVenda', backref='venda', lazy=True)
    pagamentos = db.relationship('PagamentoVenda', backref='venda', lazy=True)

class PagamentoVenda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'), nullable=False)
    tipo_pagamento = db.Column(db.String(50), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    parcelas = db.Column(db.Integer, default=1)

class ItemVenda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'), nullable=False)
    produto_id = db.Column(db.Integer, db.ForeignKey('produto.id'), nullable=False)
    quantidade = db.Column(db.Integer, default=1)
    valor_unitario = db.Column(db.Float, default=0.0)
    nome_produto = db.Column(db.String(200))  # Nome do produto no momento da venda
    produto = db.relationship('Produto')

# Novos modelos para saídas
class Devolucao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_devolucao = db.Column(db.DateTime, default=datetime.utcnow)
    valor = db.Column(db.Float, nullable=False)
    produtos_devolvidos = db.Column(db.Text)  # Lista de produtos como JSON string
    observacoes = db.Column(db.Text)
    retornar_estoque = db.Column(db.Boolean, default=False)

class PremiacaoFuncionario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_premiacao = db.Column(db.DateTime, default=datetime.utcnow)
    valor = db.Column(db.Float, nullable=False)
    funcionario = db.Column(db.String(200), nullable=False)
    descricao = db.Column(db.Text)

class AvariaProduto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_avaria = db.Column(db.DateTime, default=datetime.utcnow)
    produto_id = db.Column(db.Integer, db.ForeignKey('produto.id'), nullable=False)
    quantidade = db.Column(db.Integer, nullable=False)
    motivo = db.Column(db.String(200), nullable=False)
    observacoes = db.Column(db.Text)
    produto = db.relationship('Produto')

class CompraSuprimento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_compra = db.Column(db.DateTime, default=datetime.utcnow)
    valor = db.Column(db.Float, nullable=False)
    descricao_compra = db.Column(db.Text, nullable=False)
    fornecedor = db.Column(db.String(200))

# Modelo para controle de caixa diário
class CaixaDiario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.Date, default=date.today, unique=True)
    valor_inicial = db.Column(db.Float, default=0.0)
    valor_final = db.Column(db.Float, default=0.0)
    observacoes = db.Column(db.Text)

# Modelo para clientes/atacadistas
class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(200), nullable=False)
    tipo = db.Column(db.String(50), default='Cliente')  # 'Cliente' ou 'Atacadista'
    cpf_cnpj = db.Column(db.String(20))  # Removida constraint unique=True
    telefone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    endereco = db.Column(db.Text)
    data_cadastro = db.Column(db.DateTime, default=datetime.utcnow)
    observacoes = db.Column(db.Text)
    ativo = db.Column(db.Boolean, default=True)

# Modelo para crediário de atacadistas
class Crediario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'), nullable=False)
    valor_total = db.Column(db.Float, nullable=False)
    valor_pago = db.Column(db.Float, default=0.0)
    data_vencimento = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='Pendente')  # 'Pendente', 'Pago', 'Atrasado'
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    observacoes = db.Column(db.Text)
    cliente = db.relationship('Cliente')
    venda = db.relationship('Venda')
    pagamentos = db.relationship('PagamentoCrediario', backref='crediario', lazy=True)

class PagamentoCrediario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    crediario_id = db.Column(db.Integer, db.ForeignKey('crediario.id'), nullable=False)
    valor_pago = db.Column(db.Float, nullable=False)
    data_pagamento = db.Column(db.DateTime, default=datetime.utcnow)
    observacoes = db.Column(db.Text)

class DevolucaoCrediario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    crediario_id = db.Column(db.Integer, db.ForeignKey('crediario.id'), nullable=False)
    produtos_devolvidos = db.Column(db.Text)  # Lista de produtos como JSON string
    valor_devolvido = db.Column(db.Float, nullable=False)
    data_devolucao = db.Column(db.DateTime, default=datetime.utcnow)
    observacoes = db.Column(db.Text)
    tipo_devolucao = db.Column(db.String(20), default='caixa')  # 'caixa' ou 'credito'
    crediario = db.relationship('Crediario')

# Novo modelo para controle de crédito na loja
class CreditoLoja(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)
    origem = db.Column(db.String(100), nullable=False)  # 'devolucao_crediario', 'pagamento_adicional', etc.
    crediario_id = db.Column(db.Integer, db.ForeignKey('crediario.id'))  # Opcional, para rastrear origem
    observacoes = db.Column(db.Text)
    cliente = db.relationship('Cliente')
    crediario = db.relationship('Crediario')

# Criar banco de dados
with app.app_context():
    try:
        # Tentar criar as tabelas
        db.create_all()
        
        # Migração: adicionar coluna tipo_devolucao se não existir
        try:
            # Verificar se a coluna tipo_devolucao existe na tabela devolucao_crediario
            result = db.session.execute(text("PRAGMA table_info(devolucao_crediario)"))
            colunas = [row[1] for row in result]
            
            if 'tipo_devolucao' not in colunas:
                print("Adicionando coluna tipo_devolucao à tabela devolucao_crediario...")
                db.session.execute(text("ALTER TABLE devolucao_crediario ADD COLUMN tipo_devolucao VARCHAR(20) DEFAULT 'caixa'"))
                db.session.commit()
                print("Coluna tipo_devolucao adicionada com sucesso!")
            else:
                print("Coluna tipo_devolucao já existe na tabela devolucao_crediario.")
                
        except Exception as e:
            print(f"Erro na migração da coluna tipo_devolucao: {e}")
        
        # Migração: adicionar coluna valor_custo na tabela produto se não existir
        try:
            result_prod = db.session.execute(text("PRAGMA table_info(produto)"))
            colunas_prod = [row[1] for row in result_prod]
            if 'valor_custo' not in colunas_prod:
                print("Adicionando coluna valor_custo à tabela produto...")
                db.session.execute(text("ALTER TABLE produto ADD COLUMN valor_custo FLOAT DEFAULT 0"))
                db.session.commit()
                print("Coluna valor_custo adicionada com sucesso!")
            else:
                print("Coluna valor_custo já existe na tabela produto.")
        except Exception as e:
            print(f"Erro na migração da coluna valor_custo: {e}")
        
        print("Banco de dados criado/atualizado com sucesso!")
    except Exception as e:
        print(f"Erro ao criar banco de dados: {e}")
        # Se houver erro, tentar recriar
        try:
            db.drop_all()
            db.create_all()
            print("Banco de dados recriado com sucesso!")
        except Exception as e2:
            print(f"Erro ao recriar banco de dados: {e2}")

# Rotas principais
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/estoque')
def estoque():
    produtos = Produto.query.all()
    return render_template('estoque.html', produtos=produtos)

@app.route('/vendas')
def vendas():
    vendas = Venda.query.order_by(Venda.data_venda.desc()).all()
    produtos = Produto.query.all()
    return render_template('vendas_simples.html', vendas=vendas, produtos=produtos)

@app.route('/relatorios')
def relatorios():
    return render_template('relatorios.html')

@app.route('/saidas')
def saidas():
    produtos = Produto.query.all()
    return render_template('saidas.html', produtos=produtos)

@app.route('/teste')
def teste():
    return render_template('teste_vendas.html')

@app.route('/vendas-simples')
def vendas_simples():
    return render_template('vendas_simples.html')

@app.route('/clientes')
def clientes():
    return render_template('clientes.html')

@app.route('/vendas-atacadistas')
def vendas_atacadistas():
    return render_template('vendas_atacadistas.html')

@app.route('/pagamentos-crediarios')
def pagamentos_crediarios():
    return render_template('pagamentos_crediarios.html')

# API para produtos
@app.route('/api/produtos', methods=['GET'])
def get_produtos():
    produtos = Produto.query.all()
    return jsonify([{
        'id': p.id,
        'nome': p.nome,
        'quantidade': p.quantidade,
        'valor_custo': getattr(p, 'valor_custo', 0.0),
        'valor_varejo': p.valor_varejo,  # Novo campo
        'valor_atacado': p.valor_atacado,  # Novo campo
        'compatibilidade': p.compatibilidade,
        'total_estoque': round((getattr(p, 'valor_custo', 0.0) or 0.0) * (p.quantidade or 0), 2)
    } for p in produtos])

@app.route('/api/produtos', methods=['POST'])
def adicionar_produto():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    produto = Produto(
        nome=data['nome'],
        quantidade=data['quantidade'],
        valor_custo=data.get('valor_custo', 0.0),
        valor_varejo=data.get('valor_varejo', 0.0),  # Novo campo
        valor_atacado=data.get('valor_atacado', 0.0),  # Novo campo
        compatibilidade=data['compatibilidade']
    )
    db.session.add(produto)
    db.session.commit()
    return jsonify({'success': True, 'id': produto.id})

@app.route('/api/produtos/<int:id>', methods=['PUT'])
def atualizar_produto(id):
    produto = Produto.query.get_or_404(id)
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    produto.nome = data['nome']
    produto.quantidade = data['quantidade']
    produto.valor_custo = data.get('valor_custo', getattr(produto, 'valor_custo', 0.0))
    produto.valor_varejo = data.get('valor_varejo', produto.valor_varejo)  # Novo campo
    produto.valor_atacado = data.get('valor_atacado', produto.valor_atacado)  # Novo campo
    produto.compatibilidade = data['compatibilidade']
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/produtos/<int:id>', methods=['DELETE'])
def deletar_produto(id):
    produto = Produto.query.get_or_404(id)
    db.session.delete(produto)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/produtos/busca')
def buscar_produtos():
    termo = request.args.get('q', '').strip()
    if not termo:
        return jsonify([])
    produtos = Produto.query.filter(Produto.nome.ilike(f'%{termo}%')).all()
    return jsonify([
        {
            'id': p.id,
            'nome': p.nome,
            'quantidade': p.quantidade,
            'valor_custo': getattr(p, 'valor_custo', 0.0),
            'valor_varejo': p.valor_varejo,
            'valor_atacado': p.valor_atacado,
            'total_estoque': round((getattr(p, 'valor_custo', 0.0) or 0.0) * (p.quantidade or 0), 2)
        }
        for p in produtos
    ])

# API para clientes
@app.route('/api/clientes', methods=['GET'])
def get_clientes():
    # Verificar se há parâmetro de busca
    busca = request.args.get('busca', '').strip()
    
    if busca:
        # Buscar por CPF/CNPJ ou nome
        clientes = Cliente.query.filter(
            Cliente.ativo == True,
            db.or_(
                Cliente.cpf_cnpj.contains(busca),
                Cliente.nome.contains(busca)
            )
        ).order_by(Cliente.nome).all()
    else:
        # Buscar todos os clientes ativos
        clientes = Cliente.query.filter_by(ativo=True).order_by(Cliente.nome).all()
    
    return jsonify([{
        'id': c.id,
        'nome': c.nome,
        'tipo': c.tipo,
        'cpf_cnpj': c.cpf_cnpj,
        'telefone': c.telefone,
        'email': c.email,
        'endereco': c.endereco,
        'data_cadastro': c.data_cadastro.strftime('%d/%m/%Y'),
        'observacoes': c.observacoes
    } for c in clientes])

@app.route('/api/clientes', methods=['POST'])
def adicionar_cliente():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    # Verificar se CPF/CNPJ já existe (apenas em clientes ativos)
    if data.get('cpf_cnpj'):
        cliente_existente = Cliente.query.filter_by(cpf_cnpj=data['cpf_cnpj'], ativo=True).first()
        if cliente_existente:
            return jsonify({'success': False, 'error': 'CPF/CNPJ já cadastrado'}), 400
    
    cliente = Cliente(
        nome=data['nome'],
        tipo=data.get('tipo', 'Cliente'),
        cpf_cnpj=data.get('cpf_cnpj', ''),
        telefone=data.get('telefone', ''),
        email=data.get('email', ''),
        endereco=data.get('endereco', ''),
        observacoes=data.get('observacoes', '')
    )
    db.session.add(cliente)
    db.session.commit()
    return jsonify({'success': True, 'id': cliente.id})

@app.route('/api/clientes/<int:id>', methods=['PUT'])
def atualizar_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    # Verificar se CPF/CNPJ já existe em outro cliente ativo
    if data.get('cpf_cnpj') and data['cpf_cnpj'] != cliente.cpf_cnpj:
        cliente_existente = Cliente.query.filter_by(cpf_cnpj=data['cpf_cnpj'], ativo=True).first()
        if cliente_existente:
            return jsonify({'success': False, 'error': 'CPF/CNPJ já cadastrado'}), 400
    
    cliente.nome = data['nome']
    cliente.tipo = data.get('tipo', 'Cliente')
    cliente.cpf_cnpj = data.get('cpf_cnpj', '')
    cliente.telefone = data.get('telefone', '')
    cliente.email = data.get('email', '')
    cliente.endereco = data.get('endereco', '')
    cliente.observacoes = data.get('observacoes', '')
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/clientes/<int:id>', methods=['DELETE'])
def deletar_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    cliente.ativo = False  # Soft delete
    db.session.commit()
    return jsonify({'success': True})

# API para crediários
@app.route('/api/crediarios', methods=['GET'])
def get_crediarios():
    crediarios = Crediario.query.order_by(Crediario.data_vencimento.asc()).all()
    return jsonify([{
        'id': c.id,
        'cliente_nome': c.cliente.nome,
        'cliente_id': c.cliente_id,
        'venda_id': c.venda_id,
        'valor_total': c.venda.valor_total if c.venda else c.valor_total,  # Sempre o valor original da venda
        'valor_atual': c.valor_total,  # Valor atual após devoluções
        'valor_pago': c.valor_pago,
        'valor_restante': c.valor_total - c.valor_pago,
        'data_vencimento': c.data_vencimento.strftime('%d/%m/%Y'),
        'status': c.status,
        'data_criacao': c.data_criacao.strftime('%d/%m/%Y'),
        'observacoes': c.observacoes
    } for c in crediarios])

@app.route('/api/crediarios', methods=['POST'])
def criar_crediario():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    crediario = Crediario(
        cliente_id=data['cliente_id'],
        venda_id=data['venda_id'],
        valor_total=data['valor_total'],
        data_vencimento=datetime.strptime(data['data_vencimento'], '%Y-%m-%d').date(),
        observacoes=data.get('observacoes', '')
    )
    db.session.add(crediario)
    db.session.commit()
    return jsonify({'success': True, 'id': crediario.id})

@app.route('/api/crediarios/<int:id>/pagar', methods=['POST'])
def pagar_crediario(id):
    crediario = Crediario.query.get_or_404(id)
    data = request.json
    
    if not data or 'valor_pago' not in data:
        return jsonify({'success': False, 'error': 'Valor de pagamento é obrigatório'}), 400
    
    # Verificar se há valor restante para pagar
    valor_restante = crediario.valor_total - crediario.valor_pago
    print(f"DEBUG ORIGINAL: crediario.valor_total={crediario.valor_total}, crediario.valor_pago={crediario.valor_pago}")
    print(f"DEBUG ORIGINAL: valor_restante calculado={valor_restante}")
    
    if valor_restante <= 0:
        return jsonify({'success': False, 'error': 'Não há valor restante para pagar'}), 400
    
    valor_pago = float(data['valor_pago'])
    print(f"DEBUG ORIGINAL: valor_pago recebido={valor_pago}")
    
    # Validações
    if valor_pago <= 0:
        return jsonify({'success': False, 'error': 'Valor de pagamento deve ser maior que zero'}), 400
    
    # Arredondar valores para 2 casas decimais para evitar problemas de precisão
    valor_pago_rounded = round(valor_pago, 2)
    valor_restante_rounded = round(valor_restante, 2)
    
    print(f"DEBUG ARREDONDADO: valor_pago={valor_pago} -> {valor_pago_rounded}")
    print(f"DEBUG ARREDONDADO: valor_restante={valor_restante} -> {valor_restante_rounded}")
    print(f"DEBUG COMPARAÇÃO: {valor_pago_rounded} > {valor_restante_rounded} = {valor_pago_rounded > valor_restante_rounded}")
    
    # Comparar valores arredondados
    if valor_pago_rounded > valor_restante_rounded:
        return jsonify({
            'success': False, 
            'error': f'Valor de pagamento (R$ {valor_pago_rounded:.2f}) não pode ser maior que o valor restante (R$ {valor_restante_rounded:.2f})'
        }), 400
    
    # Registrar o pagamento
    pagamento = PagamentoCrediario(
        crediario_id=crediario.id,
        valor_pago=valor_pago_rounded,  # Usar valor arredondado
        observacoes=data.get('observacoes', '')
    )
    db.session.add(pagamento)
    
    # Atualizar valor pago do crediário
    crediario.valor_pago += valor_pago_rounded  # Usar valor arredondado
    
    # Atualizar observações se fornecidas
    if data.get('observacoes'):
        if crediario.observacoes:
            crediario.observacoes += f"\n--- Pagamento em {datetime.now().strftime('%d/%m/%Y %H:%M')} ---\n{data['observacoes']}"
        else:
            crediario.observacoes = f"Pagamento em {datetime.now().strftime('%d/%m/%Y %H:%M')}: {data['observacoes']}"
    
    # Atualizar status
    if crediario.valor_pago >= crediario.valor_total:
        crediario.status = 'Pago'
    elif crediario.data_vencimento < date.today():
        crediario.status = 'Atrasado'
    else:
        crediario.status = 'Pendente'
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'valor_pago': valor_pago_rounded,
        'valor_restante': round(crediario.valor_total - crediario.valor_pago, 2),
        'status': crediario.status,
        'pagamento_id': pagamento.id
    })

@app.route('/api/crediarios/cliente/<int:cliente_id>')
def get_crediarios_cliente(cliente_id):
    try:
        crediarios = Crediario.query.filter_by(cliente_id=cliente_id).order_by(Crediario.data_vencimento.asc()).all()
        
        resultado = []
        for c in crediarios:
            # Calcular valor pago excluindo pagamentos com crédito da loja
            pagamentos_com_credito = PagamentoCrediario.query.filter(
                PagamentoCrediario.crediario_id == c.id,
                PagamentoCrediario.observacoes.contains('Pagamento com crédito na loja')
            ).all()
            
            valor_pago_com_credito = sum(p.valor_pago for p in pagamentos_com_credito)
            valor_pago_real = c.valor_pago - valor_pago_com_credito
            
            resultado.append({
                'id': c.id,
                'venda_id': c.venda_id,
                'valor_total': c.venda.valor_total if c.venda else c.valor_total,  # Sempre o valor original da venda
                'valor_atual': c.valor_total,  # Valor atual após devoluções
                'valor_pago': valor_pago_real,  # Valor pago excluindo crédito da loja
                'valor_pago_total': c.valor_pago,  # Valor pago total (incluindo crédito)
                'valor_restante': round(c.valor_total - c.valor_pago, 2),
                'data_vencimento': c.data_vencimento.strftime('%d/%m/%Y'),
                'status': c.status,
                'data_criacao': c.data_criacao.strftime('%d/%m/%Y'),
                'observacoes': c.observacoes,
                'tem_devolucao': DevolucaoCrediario.query.filter_by(crediario_id=c.id).first() is not None,
                'completamente_devolvido': _todos_produtos_devolvidos(c),
                'produtos': [
                    {
                        'nome': item.nome_produto or (item.produto.nome if item.produto else 'Produto removido'),
                        'quantidade': item.quantidade,
                        'valor_unitario': item.valor_unitario
                    }
                    for item in c.venda.produtos_vendidos
                ] if c.venda else []
            })
        
        return jsonify(resultado)
    except Exception as e:
        print(f"Erro na API de crediários: {e}")
        # Tentar executar migração se for erro de coluna não encontrada
        if "no such column" in str(e) and "tipo_devolucao" in str(e):
            try:
                print("Tentando executar migração da coluna tipo_devolucao...")
                db.session.execute(text("ALTER TABLE devolucao_crediario ADD COLUMN tipo_devolucao VARCHAR(20) DEFAULT 'caixa'"))
                db.session.commit()
                print("Migração executada com sucesso. Tentando novamente...")
                # Tentar novamente a consulta
                crediarios = Crediario.query.filter_by(cliente_id=cliente_id).order_by(Crediario.data_vencimento.asc()).all()
                
                resultado = []
                for c in crediarios:
                    # Calcular valor pago excluindo pagamentos com crédito da loja
                    pagamentos_com_credito = PagamentoCrediario.query.filter(
                        PagamentoCrediario.crediario_id == c.id,
                        PagamentoCrediario.observacoes.contains('Pagamento com crédito na loja')
                    ).all()
                    
                    valor_pago_com_credito = sum(p.valor_pago for p in pagamentos_com_credito)
                    valor_pago_real = c.valor_pago - valor_pago_com_credito
                    
                    resultado.append({
                        'id': c.id,
                        'venda_id': c.venda_id,
                        'valor_total': c.venda.valor_total if c.venda else c.valor_total,
                        'valor_atual': c.valor_total,
                        'valor_pago': valor_pago_real,  # Valor pago excluindo crédito da loja
                        'valor_pago_total': c.valor_pago,  # Valor pago total (incluindo crédito)
                        'valor_restante': round(c.valor_total - c.valor_pago, 2),
                        'data_vencimento': c.data_vencimento.strftime('%d/%m/%Y'),
                        'status': c.status,
                        'data_criacao': c.data_criacao.strftime('%d/%m/%Y'),
                        'observacoes': c.observacoes,
                        'tem_devolucao': DevolucaoCrediario.query.filter_by(crediario_id=c.id).first() is not None,
                        'completamente_devolvido': _todos_produtos_devolvidos(c),
                        'produtos': [
                            {
                                'nome': item.nome_produto or (item.produto.nome if item.produto else 'Produto removido'),
                                'quantidade': item.quantidade,
                                'valor_unitario': item.valor_unitario
                            }
                            for item in c.venda.produtos_vendidos
                        ] if c.venda else []
                    })
                
                return jsonify(resultado)
            except Exception as e2:
                print(f"Erro na migração: {e2}")
                return jsonify({'error': 'Erro na migração do banco de dados'}), 500
        return jsonify({'error': str(e)}), 500

def _todos_produtos_devolvidos(crediario):
    """Retorna True se todos os produtos da venda do crediário foram devolvidos"""
    if not crediario.venda:
        return False
    # Soma quantidades vendidas
    vendidos = {item.produto_id: item.quantidade for item in crediario.venda.produtos_vendidos}
    # Soma quantidades devolvidas
    devolvidos = {}
    devolucoes = DevolucaoCrediario.query.filter_by(crediario_id=crediario.id).all()
    import json
    for dev in devolucoes:
        try:
            lista = json.loads(dev.produtos_devolvidos)
            for p in lista:
                devolvidos[p['id']] = devolvidos.get(p['id'], 0) + p['quantidade']
        except:
            pass
    # Verifica se todos os produtos vendidos foram totalmente devolvidos
    for pid, qtd in vendidos.items():
        if devolvidos.get(pid, 0) < qtd:
            return False
    return True

@app.route('/api/crediarios/<int:id>')
def get_crediario_detalhado(id):
    crediario = Crediario.query.get_or_404(id)
    return jsonify({
        'id': crediario.id,
        'cliente_nome': crediario.cliente.nome,
        'cliente_id': crediario.cliente_id,
        'venda_id': crediario.venda_id,
        'valor_total': crediario.venda.valor_total if crediario.venda else crediario.valor_total,  # Sempre o valor original da venda
        'valor_atual': crediario.valor_total,  # Valor atual após devoluções
        'valor_pago': crediario.valor_pago,
        'valor_restante': crediario.valor_total - crediario.valor_pago,
        'data_vencimento': crediario.data_vencimento.strftime('%d/%m/%Y'),
        'status': crediario.status,
        'data_criacao': crediario.data_criacao.strftime('%d/%m/%Y'),
        'observacoes': crediario.observacoes,
        'dias_vencimento': (crediario.data_vencimento - date.today()).days,
        'produtos': [
            {
                'nome': item.nome_produto or (item.produto.nome if item.produto else 'Produto removido'),
                'quantidade': item.quantidade,
                'valor_unitario': item.valor_unitario
            }
            for item in crediario.venda.produtos_vendidos
        ] if crediario.venda else []
    })

@app.route('/api/crediarios/<int:id>/pagamentos')
def get_pagamentos_crediario(id):
    crediario = Crediario.query.get_or_404(id)
    pagamentos = PagamentoCrediario.query.filter_by(crediario_id=id).order_by(PagamentoCrediario.data_pagamento.desc()).all()
    
    return jsonify([{
        'id': p.id,
        'valor_pago': p.valor_pago,
        'data_pagamento': p.data_pagamento.strftime('%d/%m/%Y %H:%M'),
        'observacoes': p.observacoes
    } for p in pagamentos])

@app.route('/api/crediarios/<int:id>/devolver', methods=['POST'])
def devolver_crediario(id):
    crediario = Crediario.query.get_or_404(id)
    data = request.json
    
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    produtos_devolvidos_json = data.get('produtos_devolvidos', '')
    tipo_devolucao = data.get('tipo_devolucao', 'caixa')  # 'caixa' ou 'credito'
    
    try:
        produtos_devolvidos = []
        if produtos_devolvidos_json:
            import json
            produtos_devolvidos = json.loads(produtos_devolvidos_json)
    except Exception as e:
        return jsonify({'success': False, 'error': 'Produtos devolvidos em formato inválido'}), 400

    # Calcular valor_devolvido com base nos produtos devolvidos
    valor_devolvido = 0
    for prod in produtos_devolvidos:
        valor_devolvido += prod.get('valor_unitario', 0) * prod.get('quantidade', 0)

    # Devolver produtos ao estoque
    for prod in produtos_devolvidos:
        produto = Produto.query.get(prod['id'])
        if produto:
            produto.quantidade += prod['quantidade']

    # Verificar se todos os produtos foram devolvidos após esta devolução
    def _todos_produtos_devolvidos_apos(crediario, nova_devolucao):
        if not crediario.venda:
            return False
        vendidos = {item.produto_id: item.quantidade for item in crediario.venda.produtos_vendidos}
        devolvidos = {}
        devolucoes = DevolucaoCrediario.query.filter_by(crediario_id=crediario.id).all()
        import json
        for dev in devolucoes:
            try:
                lista = json.loads(dev.produtos_devolvidos)
                for p in lista:
                    devolvidos[p['id']] = devolvidos.get(p['id'], 0) + p['quantidade']
            except:
                pass
        # Adiciona a devolução atual
        for p in nova_devolucao:
            devolvidos[p['id']] = devolvidos.get(p['id'], 0) + p['quantidade']
        for pid, qtd in vendidos.items():
            if devolvidos.get(pid, 0) < qtd:
                return False
        return True

    todos_devolvidos = _todos_produtos_devolvidos_apos(crediario, produtos_devolvidos)

    foi_pago = crediario.valor_pago > 0

    if not foi_pago:
        if todos_devolvidos:
            crediario.valor_total = 0
            crediario.status = 'Retorno'
        else:
            crediario.valor_total -= valor_devolvido
            if crediario.valor_total < 0:
                crediario.valor_total = 0
            # Status permanece Pendente
    else:
        if crediario.status == 'Pago':
            if valor_devolvido > crediario.valor_pago:
                return jsonify({'success': False, 'error': 'Valor da devolução não pode ser maior que o valor pago'}), 400
            
            # Nova lógica: verificar se deve retirar do caixa ou adicionar como crédito
            if tipo_devolucao == 'caixa':
                # Retirar do caixa diário (comportamento atual)
                pass
            elif tipo_devolucao == 'credito':
                # Adicionar como crédito na loja
                credito = CreditoLoja(
                    cliente_id=crediario.cliente_id,
                    valor=valor_devolvido,
                    origem='devolucao_crediario',
                    crediario_id=crediario.id,
                    observacoes=f"Crédito gerado por devolução de produtos pagos - Crediário ID {id}"
                )
                db.session.add(credito)
            # Status permanece 'Pago'
        else:
            valor_restante = crediario.valor_total - crediario.valor_pago
            if valor_devolvido > valor_restante:
                return jsonify({'success': False, 'error': 'Valor da devolução não pode ser maior que o valor restante'}), 400
            crediario.valor_total -= valor_devolvido
            if crediario.valor_total < 0:
                crediario.valor_total = 0
            if todos_devolvidos:
                crediario.status = 'Retorno'
            elif crediario.valor_pago >= crediario.valor_total:
                crediario.status = 'Pago'
            elif crediario.data_vencimento < date.today():
                crediario.status = 'Atrasado'
            else:
                crediario.status = 'Pendente'

    # Registrar a devolução de crediário
    devolucao = DevolucaoCrediario(
        crediario_id=id,
        produtos_devolvidos=produtos_devolvidos_json,
        valor_devolvido=valor_devolvido,
        observacoes=data.get('observacoes', ''),
        tipo_devolucao=tipo_devolucao
    )
    db.session.add(devolucao)
    
    # Registrar no histórico geral de devoluções (saídas) apenas se for tipo 'caixa'
    if tipo_devolucao == 'caixa':
        devolucao_saida = Devolucao(
            valor=valor_devolvido,
            produtos_devolvidos=produtos_devolvidos_json,
            observacoes=f"[DEVOLUÇÃO ATACADISTA - Crediário ID {id}] " + data.get('observacoes', '')
        )
        db.session.add(devolucao_saida)
    
    # Adicionar observação sobre a devolução
    if crediario.observacoes:
        crediario.observacoes += f"\n--- Devolução em {datetime.now().strftime('%d/%m/%Y %H:%M')} ({tipo_devolucao}) ---\n{data.get('observacoes', '')}"
    else:
        crediario.observacoes = f"Devolução em {datetime.now().strftime('%d/%m/%Y %H:%M')} ({tipo_devolucao}): {data.get('observacoes', '')}"
    
    db.session.commit()
    return jsonify({
        'success': True,
        'valor_restante': crediario.valor_total - crediario.valor_pago,
        'status': crediario.status,
        'devolucao_id': devolucao.id,
        'tipo_devolucao': tipo_devolucao
    })

@app.route('/api/crediarios/<int:id>/devolucoes')
def get_devolucoes_crediario(id):
    try:
        devolucoes = DevolucaoCrediario.query.filter_by(crediario_id=id).order_by(DevolucaoCrediario.data_devolucao.desc()).all()
        
        return jsonify([{
            'id': d.id,
            'valor_devolvido': d.valor_devolvido,
            'data_devolucao': d.data_devolucao.strftime('%d/%m/%Y %H:%M'),
            'produtos_devolvidos': d.produtos_devolvidos,
            'observacoes': d.observacoes,
            'tipo_devolucao': getattr(d, 'tipo_devolucao', 'caixa')  # Usar 'caixa' como padrão se a coluna não existir
        } for d in devolucoes])
    except Exception as e:
        print(f"Erro ao buscar devoluções do crediário: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/crediarios/cliente/<int:cliente_id>/devolucoes')
def get_devolucoes_cliente(cliente_id):
    try:
        crediarios = Crediario.query.filter_by(cliente_id=cliente_id).all()
        devolucoes = []
        
        for crediario in crediarios:
            devolucoes_crediario = DevolucaoCrediario.query.filter_by(crediario_id=crediario.id).all()
            for devolucao in devolucoes_crediario:
                devolucoes.append({
                    'id': devolucao.id,
                    'crediario_id': crediario.id,
                    'valor_devolvido': devolucao.valor_devolvido,
                    'data_devolucao': devolucao.data_devolucao.strftime('%d/%m/%Y %H:%M'),
                    'produtos_devolvidos': devolucao.produtos_devolvidos,
                    'observacoes': devolucao.observacoes,
                    'tipo_devolucao': getattr(devolucao, 'tipo_devolucao', 'caixa')  # Usar 'caixa' como padrão se a coluna não existir
                })
        
        return jsonify(devolucoes)
    except Exception as e:
        print(f"Erro ao buscar devoluções do cliente: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/crediarios/<int:id>/produtos-disponiveis-devolucao')
def get_produtos_disponiveis_devolucao(id):
    crediario = Crediario.query.get_or_404(id)
    
    # Verificar se já existe devolução
    devolucao_existente = DevolucaoCrediario.query.filter_by(crediario_id=id).first()
    if devolucao_existente:
        return jsonify([])  # Não há produtos disponíveis se já foi devolvido
    
    # Retornar produtos da venda
    produtos = []
    if crediario.venda and crediario.venda.produtos_vendidos:
        for item in crediario.venda.produtos_vendidos:
            produtos.append({
                'id': item.produto_id,
                'nome': item.nome_produto or (item.produto.nome if item.produto else 'Produto removido'),
                'quantidade': item.quantidade,
                'valor_unitario': item.valor_unitario
            })
    
    return jsonify(produtos)

# APIs para Crédito na Loja
@app.route('/api/credito-loja/cliente/<int:cliente_id>')
def get_credito_loja_cliente(cliente_id):
    """Retorna o saldo de crédito na loja de um cliente"""
    creditos = CreditoLoja.query.filter_by(cliente_id=cliente_id).all()
    total_credito = sum(credito.valor for credito in creditos)
    
    return jsonify({
        'cliente_id': cliente_id,
        'total_credito': total_credito,
        'creditos': [{
            'id': c.id,
            'valor': c.valor,
            'data_criacao': c.data_criacao.strftime('%d/%m/%Y %H:%M'),
            'origem': c.origem,
            'crediario_id': c.crediario_id,
            'observacoes': c.observacoes
        } for c in creditos]
    })

@app.route('/api/credito-loja/cliente/<int:cliente_id>/usar', methods=['POST'])
def usar_credito_loja(cliente_id):
    """Usa crédito da loja para pagar um crediário"""
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    valor_usar = float(data.get('valor', 0))
    crediario_id = data.get('crediario_id')
    observacoes = data.get('observacoes', '')
    
    if valor_usar <= 0:
        return jsonify({'success': False, 'error': 'Valor deve ser maior que zero'}), 400
    
    # Verificar se o cliente tem crédito suficiente
    creditos = CreditoLoja.query.filter_by(cliente_id=cliente_id).all()
    total_credito = sum(credito.valor for credito in creditos)
    
    if valor_usar > total_credito:
        return jsonify({'success': False, 'error': 'Crédito insuficiente na loja'}), 400
    
    # Se foi especificado um crediário, usar o crédito para pagá-lo
    if crediario_id:
        crediario = Crediario.query.get_or_404(crediario_id)
        if crediario.cliente_id != cliente_id:
            return jsonify({'success': False, 'error': 'Crediário não pertence ao cliente'}), 400
        
        valor_restante_crediario = crediario.valor_total - crediario.valor_pago
        if valor_usar > valor_restante_crediario:
            return jsonify({'success': False, 'error': 'Valor maior que o valor restante do crediário'}), 400
        
        # Registrar pagamento do crediário
        pagamento = PagamentoCrediario(
            crediario_id=crediario_id,
            valor_pago=valor_usar,
            observacoes=f"Pagamento com crédito na loja: {observacoes}"
        )
        db.session.add(pagamento)
        
        # Atualizar valor pago do crediário
        crediario.valor_pago += valor_usar
        
        # Atualizar status do crediário
        if crediario.valor_pago >= crediario.valor_total:
            crediario.status = 'Pago'
        elif crediario.data_vencimento < date.today():
            crediario.status = 'Atrasado'
        else:
            crediario.status = 'Pendente'
    
    # Criar registro de uso do crédito (valor negativo)
    uso_credito = CreditoLoja(
        cliente_id=cliente_id,
        valor=-valor_usar,
        origem='uso_credito',
        crediario_id=crediario_id,
        observacoes=f"Uso de crédito na loja: {observacoes}"
    )
    db.session.add(uso_credito)
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'valor_usado': valor_usar,
        'credito_restante': total_credito - valor_usar
    })

@app.route('/api/credito-loja/cliente/<int:cliente_id>/adicionar', methods=['POST'])
def adicionar_credito_loja(cliente_id):
    """Adiciona crédito na loja para um cliente"""
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    valor = float(data.get('valor', 0))
    origem = data.get('origem', 'pagamento_adicional')
    observacoes = data.get('observacoes', '')
    
    if valor <= 0:
        return jsonify({'success': False, 'error': 'Valor deve ser maior que zero'}), 400
    
    # Criar registro de crédito
    credito = CreditoLoja(
        cliente_id=cliente_id,
        valor=valor,
        origem=origem,
        observacoes=observacoes
    )
    db.session.add(credito)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'id': credito.id,
        'valor_adicionado': valor
    })

# API para vendas
@app.route('/api/vendas', methods=['POST'])
def registrar_venda():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    tipo_venda = data.get('tipo_venda', 'simples')  # Novo campo para identificar o tipo de venda
    
    # Validar se a soma dos pagamentos é igual ao valor total
    valor_total_pagamentos = sum(pagamento['valor'] for pagamento in data.get('pagamentos', []))
    if abs(valor_total_pagamentos - data['valor_total']) > 0.01:  # Tolerância para diferenças de ponto flutuante
        return jsonify({'success': False, 'error': 'A soma dos valores dos pagamentos deve ser igual ao valor total da venda'}), 400
    
    # Criar venda
    venda = Venda(
        valor_total=data['valor_total'],
        parcelas=data.get('parcelas', 1),
        cliente=data.get('cliente', '')
    )
    db.session.add(venda)
    db.session.flush()  # Para obter o ID da venda
    
    # Adicionar pagamentos da venda
    for pagamento_data in data.get('pagamentos', []):
        pagamento = PagamentoVenda(
            venda_id=venda.id,
            tipo_pagamento=pagamento_data['tipo_pagamento'],
            valor=pagamento_data['valor'],
            parcelas=pagamento_data.get('parcelas', 1)
        )
        db.session.add(pagamento)
    
    # Adicionar itens da venda e atualizar estoque
    for item_data in data['itens']:
        produto = Produto.query.get(item_data['produto_id'])
        if produto and produto.quantidade >= item_data['quantidade']:
            # Definir preço conforme o tipo de venda
            if tipo_venda == 'atacadista':
                valor_unitario = produto.valor_atacado
            else:
                valor_unitario = produto.valor_varejo
            # Criar item da venda
            item_venda = ItemVenda(
                venda_id=venda.id,
                produto_id=produto.id,
                quantidade=item_data['quantidade'],
                valor_unitario=valor_unitario,
                nome_produto=produto.nome  # Salvar o nome do produto no momento da venda
            )
            db.session.add(item_venda)
            
            # Atualizar estoque
            produto.quantidade -= item_data['quantidade']
        else:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'Estoque insuficiente'})
    
    db.session.commit()
    return jsonify({'success': True, 'id': venda.id})

@app.route('/api/vendas')
def get_vendas():
    vendas = Venda.query.order_by(Venda.data_venda.desc()).all()
    return jsonify([{
        'id': v.id,
        'data_venda': v.data_venda.strftime('%d/%m/%Y %H:%M'),
        'valor_total': v.valor_total,
        'parcelas': v.parcelas,
        'cliente': v.cliente,
        'produtos': [{
            'nome': iv.nome_produto or (iv.produto.nome if iv.produto else 'Produto removido'),
            'quantidade': iv.quantidade,
            'valor_unitario': iv.valor_unitario,
            'valor_custo': getattr(iv.produto, 'valor_custo', 0.0) if iv.produto else 0.0
        } for iv in v.produtos_vendidos],
        'pagamentos': [{
            'tipo_pagamento': p.tipo_pagamento,
            'valor': p.valor,
            'parcelas': p.parcelas
        } for p in v.pagamentos],
        'lucro': sum([
            (iv.valor_unitario - (getattr(iv.produto, 'valor_custo', 0.0) if iv.produto else 0.0)) * iv.quantidade
            for iv in v.produtos_vendidos
        ])
    } for v in vendas])

@app.route('/api/vendas/produtos/<int:venda_id>')
def get_produtos_venda(venda_id):
    venda = Venda.query.get_or_404(venda_id)
    produtos = []
    
    for item in venda.produtos_vendidos:
        produtos.append({
            'id': item.produto_id,
            'nome': item.nome_produto or (item.produto.nome if item.produto else 'Produto removido'),
            'quantidade': item.quantidade,
            'valor_unitario': item.valor_unitario
        })
    
    return jsonify(produtos)

# Relatório Excel
@app.route('/api/relatorio/<data>')
def gerar_relatorio(data):
    print(f"[RELATÓRIO] Iniciando geração de relatório para data: {data}")
    
    try:
        data_relatorio = datetime.strptime(data, '%Y-%m-%d').date()
        print(f"[RELATÓRIO] Data convertida: {data_relatorio}")
    except ValueError as e:
        print(f"[RELATÓRIO] Erro na conversão da data: {e}")
        return jsonify({'error': 'Data inválida'}), 400
    
    try:
        vendas = Venda.query.filter(
            db.func.date(Venda.data_venda) == data_relatorio
        ).order_by(Venda.data_venda).all()
        
        print(f"[RELATÓRIO] Encontradas {len(vendas)} vendas para a data")
        
        # Criar planilha Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is not None:
            # Usar formato de data sem caracteres especiais para o título da planilha
            titulo_planilha = f"Relatorio {data_relatorio.strftime('%d-%m-%Y')}"
            titulo_limpo = limpar_nome_planilha(titulo_planilha)
            ws.title = titulo_limpo
            print(f"[RELATÓRIO] Título da planilha definido como: '{titulo_limpo}'")
            
            # Cabeçalhos
            headers = ['Horário', 'Valor', 'Produto(s)', 'Métodos de Pagamento', 'Cliente', 'Lucro']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Dados
            row = 2
            for venda in vendas:
                try:
                    # Preparar lista de produtos
                    produtos_str = []
                    lucro_total = 0
                    for item in venda.produtos_vendidos:
                        # Usar o nome salvo no momento da venda ou tentar buscar do produto
                        nome_produto = item.nome_produto or (item.produto.nome if item.produto else 'Produto removido')
                        produtos_str.append(f"{nome_produto} ({item.quantidade})")
                        
                        # Calcular lucro do item
                        valor_custo = getattr(item.produto, 'valor_custo', 0.0) if item.produto else 0.0
                        lucro_item = (item.valor_unitario - valor_custo) * item.quantidade
                        lucro_total += lucro_item
                    
                    # Preparar métodos de pagamento
                    pagamentos_str = []
                    for pagamento in venda.pagamentos:
                        if pagamento.tipo_pagamento == 'Cartão de Crédito' and pagamento.parcelas > 1:
                            pagamentos_str.append(f"{pagamento.tipo_pagamento} ({pagamento.parcelas}x) - R$ {pagamento.valor:.2f}")
                        else:
                            pagamentos_str.append(f"{pagamento.tipo_pagamento} - R$ {pagamento.valor:.2f}")
                    
                    ws.cell(row=row, column=1, value=venda.data_venda.strftime('%d/%m/%Y %H:%M'))
                    ws.cell(row=row, column=2, value=f"R$ {venda.valor_total:.2f}")
                    ws.cell(row=row, column=3, value=", ".join(produtos_str))
                    ws.cell(row=row, column=4, value=", ".join(pagamentos_str))
                    ws.cell(row=row, column=5, value=venda.cliente or '')
                    ws.cell(row=row, column=6, value=f"R$ {lucro_total:.2f}")
                    row += 1
                except Exception as e:
                    print(f"[RELATÓRIO] Erro ao processar venda {venda.id}: {e}")
                    continue
            
            # Adicionar linha de totalização
            if row > 2:  # Se há dados no relatório
                # Calcular totais
                total_vendas = sum(v.valor_total for v in vendas)
                total_lucro = 0
                for venda in vendas:
                    for item in venda.produtos_vendidos:
                        valor_custo = getattr(item.produto, 'valor_custo', 0.0) if item.produto else 0.0
                        lucro_item = (item.valor_unitario - valor_custo) * item.quantidade
                        total_lucro += lucro_item
                
                # Linha em branco
                row += 1
                
                # Linha de totais
                ws.cell(row=row, column=1, value="TOTAIS:")
                ws.cell(row=row, column=2, value=f"R$ {total_vendas:.2f}")
                ws.cell(row=row, column=6, value=f"R$ {total_lucro:.2f}")
                
                # Formatar linha de totais
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                
                print(f"[RELATÓRIO] Total de vendas: R$ {total_vendas:.2f}")
                print(f"[RELATÓRIO] Total de lucro: R$ {total_lucro:.2f}")
            
            # Ajustar largura das colunas
            for i, column in enumerate(ws.columns):
                max_length = 0
                column_letter = get_column_letter(i + 1)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"[RELATÓRIO] Planilha criada com {row-2} linhas de dados")
        
        # Salvar em buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"[RELATÓRIO] Arquivo salvo no buffer, tamanho: {len(output.getvalue())} bytes")
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"relatorio_vendas_{data_relatorio.strftime('%d_%m_%Y')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"[RELATÓRIO] Erro ao gerar relatório: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': 'Erro interno ao gerar relatório'}), 500

# APIs para Saídas

@app.route('/api/devolucoes', methods=['POST'])
def registrar_devolucao():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    devolucao = Devolucao(
        valor=data['valor'],
        produtos_devolvidos=data.get('produtos_devolvidos', ''),
        observacoes=data.get('observacoes', ''),
        retornar_estoque=data.get('retornar_estoque', False)
    )
    
    db.session.add(devolucao)
    db.session.commit()
    return jsonify({'success': True, 'id': devolucao.id})

@app.route('/api/devolucoes')
def get_devolucoes():
    devolucoes = Devolucao.query.order_by(Devolucao.data_devolucao.desc()).all()
    return jsonify([{
        'id': d.id,
        'data_devolucao': d.data_devolucao.strftime('%d/%m/%Y %H:%M'),
        'valor': d.valor,
        'produtos_devolvidos': d.produtos_devolvidos,
        'observacoes': d.observacoes,
        'retornar_estoque': d.retornar_estoque
    } for d in devolucoes])

@app.route('/api/premiacoes', methods=['POST'])
def registrar_premiacao():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    premiacao = PremiacaoFuncionario(
        valor=data['valor'],
        funcionario=data['funcionario'],
        descricao=data.get('descricao', '')
    )
    
    db.session.add(premiacao)
    db.session.commit()
    return jsonify({'success': True, 'id': premiacao.id})

@app.route('/api/premiacoes')
def get_premiacoes():
    premiacoes = PremiacaoFuncionario.query.order_by(PremiacaoFuncionario.data_premiacao.desc()).all()
    return jsonify([{
        'id': p.id,
        'data_premiacao': p.data_premiacao.strftime('%d/%m/%Y %H:%M'),
        'valor': p.valor,
        'funcionario': p.funcionario,
        'descricao': p.descricao
    } for p in premiacoes])

@app.route('/api/avarias', methods=['POST'])
def registrar_avaria():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    produto = Produto.query.get(data['produto_id'])
    if not produto:
        return jsonify({'success': False, 'error': 'Produto não encontrado'}), 404
    
    if produto.quantidade < data['quantidade']:
        return jsonify({'success': False, 'error': 'Quantidade insuficiente em estoque'}), 400
    
    avaria = AvariaProduto(
        produto_id=data['produto_id'],
        quantidade=data['quantidade'],
        motivo=data['motivo'],
        observacoes=data.get('observacoes', '')
    )
    
    # Atualizar estoque
    produto.quantidade -= data['quantidade']
    
    db.session.add(avaria)
    db.session.commit()
    return jsonify({'success': True, 'id': avaria.id})

@app.route('/api/avarias')
def get_avarias():
    avarias = AvariaProduto.query.order_by(AvariaProduto.data_avaria.desc()).all()
    return jsonify([{
        'id': a.id,
        'data_avaria': a.data_avaria.strftime('%d/%m/%Y %H:%M'),
        'produto_nome': a.produto.nome,
        'quantidade': a.quantidade,
        'motivo': a.motivo,
        'observacoes': a.observacoes
    } for a in avarias])

@app.route('/api/compras-suprimentos', methods=['POST'])
def registrar_compra_suprimento():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    compra = CompraSuprimento(
        valor=data['valor'],
        descricao_compra=data['descricao_compra'],
        fornecedor=data.get('fornecedor', '')
    )
    
    db.session.add(compra)
    db.session.commit()
    return jsonify({'success': True, 'id': compra.id})

@app.route('/api/compras-suprimentos')
def get_compras_suprimentos():
    compras = CompraSuprimento.query.order_by(CompraSuprimento.data_compra.desc()).all()
    return jsonify([{
        'id': c.id,
        'data_compra': c.data_compra.strftime('%d/%m/%Y %H:%M'),
        'valor': c.valor,
        'descricao_compra': c.descricao_compra,
        'fornecedor': c.fornecedor
    } for c in compras])

# APIs para Controle de Caixa Diário

def calcular_caixa_diario(data_caixa=None):
    """Calcula o saldo do caixa para uma data específica"""
    if data_caixa is None:
        data_caixa = date.today()
    
    # Buscar ou criar registro do caixa diário
    caixa = CaixaDiario.query.filter_by(data=data_caixa).first()
    if not caixa:
        caixa = CaixaDiario(data=data_caixa, valor_inicial=0.0, valor_final=0.0)
        db.session.add(caixa)
        db.session.commit()
    
    # IDs de vendas que são crediário (garantir set para robustez)
    vendas_crediario_ids = set(c.venda_id for c in Crediario.query.all())
    # Vendas do dia
    vendas_do_dia = Venda.query.filter(
        db.func.date(Venda.data_venda) == data_caixa
    ).all()
    vendas_do_dia_ids = [v.id for v in vendas_do_dia]
    vendas_simples_ids = [v.id for v in vendas_do_dia if v.id not in vendas_crediario_ids]
    vendas_crediario_do_dia_ids = [v.id for v in vendas_do_dia if v.id in vendas_crediario_ids]
    print(f"[CAIXA] Data: {data_caixa}")
    print(f"[CAIXA] Vendas do dia: {vendas_do_dia_ids}")
    print(f"[CAIXA] Vendas com crediário: {list(vendas_crediario_ids)}")
    print(f"[CAIXA] Vendas simples (entram no caixa): {vendas_simples_ids}")
    print(f"[CAIXA] Vendas crediário (NÃO entram no caixa): {vendas_crediario_do_dia_ids}")
    # Somar apenas vendas que NÃO estão no set de crediário
    total_vendas_simples = sum(venda.valor_total for venda in vendas_do_dia if venda.id not in vendas_crediario_ids)
    
    # Pagamentos de crediário recebidos no dia (excluindo pagamentos com crédito da loja)
    pagamentos_crediario = PagamentoCrediario.query.filter(
        db.func.date(PagamentoCrediario.data_pagamento) == data_caixa,
        ~PagamentoCrediario.observacoes.contains('Pagamento com crédito na loja')
    ).all()
    total_pagamentos_crediario = sum(p.valor_pago for p in pagamentos_crediario)
    
    # Log para debug
    pagamentos_com_credito = PagamentoCrediario.query.filter(
        db.func.date(PagamentoCrediario.data_pagamento) == data_caixa,
        PagamentoCrediario.observacoes.contains('Pagamento com crédito na loja')
    ).all()
    total_pagamentos_com_credito = sum(p.valor_pago for p in pagamentos_com_credito)
    
    print(f"[CAIXA] Pagamentos crediário (entram no caixa): {total_pagamentos_crediario}")
    print(f"[CAIXA] Pagamentos com crédito da loja (NÃO entram no caixa): {total_pagamentos_com_credito}")
    
    # Calcular saídas (devoluções, premiações, compras)
    devolucoes_do_dia = Devolucao.query.filter(
        db.func.date(Devolucao.data_devolucao) == data_caixa
    ).all()
    total_devolucoes = sum(devolucao.valor for devolucao in devolucoes_do_dia)
    
    # Calcular devoluções de crediário que afetam o caixa (apenas tipo 'caixa')
    devolucoes_crediario_caixa = DevolucaoCrediario.query.filter(
        db.func.date(DevolucaoCrediario.data_devolucao) == data_caixa,
        DevolucaoCrediario.tipo_devolucao == 'caixa'
    ).all()
    total_devolucoes_crediario_caixa = sum(devolucao.valor_devolvido for devolucao in devolucoes_crediario_caixa)
    
    # Total de devoluções (gerais + crediário tipo caixa)
    total_devolucoes += total_devolucoes_crediario_caixa
    
    premiacoes_do_dia = PremiacaoFuncionario.query.filter(
        db.func.date(PremiacaoFuncionario.data_premiacao) == data_caixa
    ).all()
    total_premiacoes = sum(premiacao.valor for premiacao in premiacoes_do_dia)
    
    compras_do_dia = CompraSuprimento.query.filter(
        db.func.date(CompraSuprimento.data_compra) == data_caixa
    ).all()
    total_compras = sum(compra.valor for compra in compras_do_dia)
    
    # Calcular saldo final
    saldo_final = caixa.valor_inicial + total_vendas_simples + total_pagamentos_crediario - total_devolucoes - total_premiacoes - total_compras
    caixa.valor_final = saldo_final
    db.session.commit()
    
    return {
        'valor_inicial': caixa.valor_inicial,
        'total_vendas': total_vendas_simples + total_pagamentos_crediario,
        'total_vendas_simples': total_vendas_simples,
        'total_pagamentos_crediario': total_pagamentos_crediario,
        'total_devolucoes': total_devolucoes,
        'total_premiacoes': total_premiacoes,
        'total_compras': total_compras,
        'saldo_final': saldo_final
    }

@app.route('/api/caixa-diario/<data>')
def get_caixa_diario(data):
    try:
        data_caixa = datetime.strptime(data, '%Y-%m-%d').date()
    except:
        return jsonify({'error': 'Data inválida'})
    
    caixa_info = calcular_caixa_diario(data_caixa)
    return jsonify(caixa_info)

@app.route('/api/caixa-diario/hoje')
def get_caixa_hoje():
    caixa_info = calcular_caixa_diario()
    return jsonify(caixa_info)

@app.route('/api/caixa-diario/valor-inicial', methods=['POST'])
def definir_valor_inicial():
    data = request.json
    if not data:
        return jsonify({'success': False, 'error': 'Dados inválidos'}), 400
    
    data_caixa = date.today()
    caixa = CaixaDiario.query.filter_by(data=data_caixa).first()
    if not caixa:
        caixa = CaixaDiario(data=data_caixa)
        db.session.add(caixa)
    
    caixa.valor_inicial = data['valor_inicial']
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/estatisticas/valor-total-estoque')
def get_valor_total_estoque():
    """Retorna o valor total do estoque baseado no valor de custo"""
    produtos = Produto.query.all()
    valor_total = sum((getattr(p, 'valor_custo', 0.0) or 0.0) * (p.quantidade or 0) for p in produtos)
    return jsonify({'valor_total_estoque': round(valor_total, 2)})

@app.route('/api/estatisticas/faturamento-liquido-diario')
def get_faturamento_liquido_diario():
    """Retorna o faturamento líquido do dia (receita - custos)"""
    data_hoje = date.today()
    
    # Buscar vendas do dia
    vendas_do_dia = Venda.query.filter(
        db.func.date(Venda.data_venda) == data_hoje
    ).all()
    
    # IDs de vendas que são crediário (não entram no caixa diário)
    vendas_crediario_ids = set(c.venda_id for c in Crediario.query.all())
    
    # Calcular receita total do dia (apenas vendas simples)
    receita_total = sum(venda.valor_total for venda in vendas_do_dia if venda.id not in vendas_crediario_ids)
    
    # Adicionar pagamentos de crediário recebidos hoje
    pagamentos_crediario = PagamentoCrediario.query.filter(
        db.func.date(PagamentoCrediario.data_pagamento) == data_hoje,
        ~PagamentoCrediario.observacoes.contains('Pagamento com crédito na loja')
    ).all()
    receita_total += sum(p.valor_pago for p in pagamentos_crediario)
    
    # Calcular custos dos produtos vendidos hoje
    custos_total = 0
    for venda in vendas_do_dia:
        for item in venda.produtos_vendidos:
            produto = Produto.query.get(item.produto_id)
            if produto:
                valor_custo = getattr(produto, 'valor_custo', 0.0) or 0.0
                custos_total += valor_custo * item.quantidade
    
    # Calcular faturamento líquido
    faturamento_liquido = receita_total - custos_total
    
    return jsonify({
        'faturamento_liquido': round(faturamento_liquido, 2),
        'receita_total': round(receita_total, 2),
        'custos_total': round(custos_total, 2)
    })

if __name__ == '__main__':
    app.run(debug=True) 